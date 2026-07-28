"""
Staff Attendance — API routes.
All endpoints are school_id scoped via the logged-in user's own school_id
(never trust a school_id from the request body/query for writes).
"""

import calendar
import io
from collections import defaultdict
from datetime import datetime, date, timedelta
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity

from app import db
from app.models.user import User, UserRole
from app.models.staff_attendance import (
    StaffAttendance, StaffAttendanceSettings, StaffAttendanceRegularization,
    StaffAttendanceAuditLog, StaffMonthlyAttendanceSummary, generate_employee_id,
)
from app.services import staff_attendance_service as svc
from app.utils.decorators import role_required

staff_attendance_bp = Blueprint('staff_attendance', __name__)


def _current_user():
    user_id = int(get_jwt_identity())
    return User.query.get(user_id)


def _parse_date(value, default=None):
    if not value:
        return default
    return datetime.strptime(value, '%Y-%m-%d').date()

_WEEKDAY_ABBR = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']


def _working_days_set(settings):
    return set((settings.working_days or '').split(','))


def _count_working_days(start, end, working_days):
    if start > end:
        return 0
    count, d = 0, start
    while d <= end:
        if _WEEKDAY_ABBR[d.weekday()] in working_days:
            count += 1
        d += timedelta(days=1)
    return count


def _build_buckets(range_key, month, year):
    """(label, start_date, end_date) tuples for the trend chart — clipped to today."""
    today = date.today()
    buckets = []
    if range_key == 'daily':
        start = today - timedelta(days=13)
        for i in range(14):
            d = start + timedelta(days=i)
            buckets.append((d.strftime('%d %b'), d, d))
    elif range_key == 'weekly':
        this_monday = today - timedelta(days=today.weekday())
        for i in range(7, -1, -1):
            w_start = this_monday - timedelta(weeks=i)
            w_end = min(w_start + timedelta(days=6), today)
            buckets.append((w_start.strftime('%d %b'), w_start, w_end))
    elif range_key == 'yearly':
        for m in range(1, 13):
            last_day = calendar.monthrange(year, m)[1]
            m_start, m_end = date(year, m, 1), date(year, m, last_day)
            if m_start > today:
                break
            buckets.append((calendar.month_abbr[m], m_start, min(m_end, today)))
    else:  # monthly — day-by-day trend within the selected month
        last_day = calendar.monthrange(year, month)[1]
        for d in range(1, last_day + 1):
            day = date(year, month, d)
            if day > today:
                break
            buckets.append((str(d), day, day))
    return buckets


def _analytics_payload(school_id, range_key, month, year, role_filter):
    """
    Single source of truth for both /analytics (JSON) and /analytics/export
    (Excel/PDF) — computed fresh each call. Given the scale target (10L+
    records), if this becomes a bottleneck later, wrap with a short-TTL
    cache keyed on (school_id, range_key, month, year, role_filter) —
    intentionally not cached yet since Settings/Leave changes should
    reflect immediately.
    """
    settings = StaffAttendanceSettings.get_or_create(school_id)
    working_days = _working_days_set(settings)
    buckets = _build_buckets(range_key, month, year)
    window_start = buckets[0][1] if buckets else date.today()
    window_end = buckets[-1][2] if buckets else date.today()

    emp_q = User.query.filter(
        User.school_id == school_id,
        User.role != UserRole.STUDENT, User.role != UserRole.PARENT,
        User.is_active == True,
    )
    if role_filter and role_filter != 'ALL':
        emp_q = emp_q.filter(User.role == role_filter)
    employees = emp_q.all()
    total_employees = len(employees)

    row_q = StaffAttendance.query.filter(
        StaffAttendance.school_id == school_id,
        StaffAttendance.attendance_date >= window_start,
        StaffAttendance.attendance_date <= window_end,
    ).join(User, StaffAttendance.user_id == User.id)
    if role_filter and role_filter != 'ALL':
        row_q = row_q.filter(User.role == role_filter)
    rows = row_q.all()

    regularizations = StaffAttendanceRegularization.query.filter(
        StaffAttendanceRegularization.school_id == school_id,
        StaffAttendanceRegularization.attendance_date >= window_start,
        StaffAttendanceRegularization.attendance_date <= window_end,
    ).all()

    # ── trend buckets ──
    trend_attendance, trend_late, trend_hours, trend_reg = [], [], [], []
    total_working_days_window = _count_working_days(window_start, window_end, working_days)

    for label, b_start, b_end in buckets:
        b_rows = [r for r in rows if b_start <= r.attendance_date <= b_end]
        expected = total_employees * _count_working_days(b_start, b_end, working_days)
        present_like = sum(1 for r in b_rows if r.status in ('PRESENT', 'LATE', 'HALF_DAY'))
        late_ct = sum(1 for r in b_rows if r.status == 'LATE')
        checked_out = [r for r in b_rows if r.check_out_time]
        avg_hrs = round((sum(r.working_minutes or 0 for r in checked_out) / len(checked_out)) / 60, 2) if checked_out else 0
        reg_ct = sum(1 for rg in regularizations if b_start <= rg.attendance_date <= b_end)

        trend_attendance.append({'label': label, 'value': round((present_like / expected) * 100, 1) if expected else 0})
        trend_late.append({'label': label, 'value': late_ct})
        trend_hours.append({'label': label, 'value': avg_hrs})
        trend_reg.append({'label': label, 'value': reg_ct})

    # ── per-employee aggregation (KPIs + rankings) ──
    per_user = defaultdict(lambda: {'attendance_days': 0, 'late_days': 0, 'half_days': 0})
    for r in rows:
        u = per_user[r.user_id]
        if r.status in ('PRESENT', 'LATE', 'HALF_DAY'):
            u['attendance_days'] += 1
        if r.status == 'LATE':
            u['late_days'] += 1
        if r.status == 'HALF_DAY':
            u['half_days'] += 1

    rankings_rows = []
    for e in employees:
        stat = per_user.get(e.id, {'attendance_days': 0, 'late_days': 0, 'half_days': 0})
        absent_days = max(0, total_working_days_window - stat['attendance_days'])
        pct = round((stat['attendance_days'] / total_working_days_window) * 100, 1) if total_working_days_window else 0
        rankings_rows.append({
            'user_id': e.id, 'employee_id': e.employee_id, 'employee_name': e.name,
            'role': e.role.value, 'designation': e.designation,
            'attendance_percent': pct, 'late_days': stat['late_days'],
            'leave_days': absent_days,  # absence-based until Leave module is wired in
        })

    top_attendance = sorted(rankings_rows, key=lambda x: x['attendance_percent'], reverse=True)[:5]
    most_late = sorted([r for r in rankings_rows if r['late_days'] > 0], key=lambda x: x['late_days'], reverse=True)[:5]
    most_leave = sorted([r for r in rankings_rows if r['leave_days'] > 0], key=lambda x: x['leave_days'], reverse=True)[:5]

    sum_attendance = sum(u['attendance_days'] for u in per_user.values())
    sum_late = sum(u['late_days'] for u in per_user.values())
    sum_half = sum(u['half_days'] for u in per_user.values())
    expected_total = total_employees * total_working_days_window
    sum_absent = max(0, expected_total - sum_attendance)

    def _avg_time(times):
        if not times:
            return None
        avg_sec = sum(t.hour * 3600 + t.minute * 60 + t.second for t in times) / len(times)
        return f'{int(avg_sec // 3600):02d}:{int((avg_sec % 3600) // 60):02d}'

    checked_out_rows = [r for r in rows if r.check_out_time]
    avg_working_hours = round((sum(r.working_minutes or 0 for r in checked_out_rows) / len(checked_out_rows)) / 60, 2) if checked_out_rows else 0

    payroll_impact = 0
    if range_key == 'monthly':
        summaries = svc.rebuild_monthly_summary_for_school(school_id, month, year)
        payroll_impact = round(sum(s.salary_impact or 0 for s in summaries), 2)

    kpis = {
        'attendance_percent': round((sum_attendance / expected_total) * 100, 1) if expected_total else 0,
        'late_percent': round((sum_late / expected_total) * 100, 1) if expected_total else 0,
        'absent_percent': round((sum_absent / expected_total) * 100, 1) if expected_total else 0,
        'half_day_percent': round((sum_half / expected_total) * 100, 1) if expected_total else 0,
        'avg_check_in': _avg_time([r.check_in_time for r in rows if r.check_in_time]),
        'avg_check_out': _avg_time([r.check_out_time for r in rows if r.check_out_time]),
        'avg_working_hours': avg_working_hours,
        'payroll_impact': payroll_impact,
    }

    # ── heatmap — always the selected calendar month, independent of `range` ──
    heat_last_day = calendar.monthrange(year, month)[1]
    month_start, month_end = date(year, month, 1), date(year, month, heat_last_day)
    month_rows = StaffAttendance.query.filter(
        StaffAttendance.school_id == school_id,
        StaffAttendance.attendance_date >= month_start,
        StaffAttendance.attendance_date <= month_end,
    ).all()
    rows_by_date = defaultdict(list)
    for r in month_rows:
        rows_by_date[r.attendance_date].append(r)

    heatmap = []
    for d in range(1, heat_last_day + 1):
        day = date(year, month, d)
        if day > date.today() or _WEEKDAY_ABBR[day.weekday()] not in working_days:
            heatmap.append({'date': day.isoformat(), 'attendance_percent': None})
            continue
        day_rows = rows_by_date.get(day, [])
        present_like = sum(1 for r in day_rows if r.status in ('PRESENT', 'LATE', 'HALF_DAY'))
        pct = round((present_like / total_employees) * 100, 1) if total_employees else None
        heatmap.append({'date': day.isoformat(), 'attendance_percent': pct})

    # ── comparison: last 6 calendar months + Apr–Mar academic session ──
    monthly_comparison = []
    for i in range(5, -1, -1):
        y, m = year, month - i
        while m <= 0:
            m += 12
            y -= 1
        if date(y, m, 1) > date.today():
            monthly_comparison.append({'label': calendar.month_abbr[m], 'value': 0})
            continue
        m_last = calendar.monthrange(y, m)[1]
        m_start, m_end = date(y, m, 1), min(date(y, m, m_last), date.today())
        m_present = StaffAttendance.query.filter(
            StaffAttendance.school_id == school_id,
            StaffAttendance.attendance_date >= m_start, StaffAttendance.attendance_date <= m_end,
            StaffAttendance.status.in_(['PRESENT', 'LATE', 'HALF_DAY']),
        ).count()
        m_expected = total_employees * _count_working_days(m_start, m_end, working_days)
        monthly_comparison.append({'label': calendar.month_abbr[m], 'value': round((m_present / m_expected) * 100, 1) if m_expected else 0})

    def _session_bounds(d):
        start_year = d.year if d.month >= 4 else d.year - 1
        return date(start_year, 4, 1), date(start_year + 1, 3, 31)

    def _session_pct(s_start, s_end):
        s_end = min(s_end, date.today())
        if s_start > s_end:
            return 0
        present_ct = StaffAttendance.query.filter(
            StaffAttendance.school_id == school_id,
            StaffAttendance.attendance_date >= s_start, StaffAttendance.attendance_date <= s_end,
            StaffAttendance.status.in_(['PRESENT', 'LATE', 'HALF_DAY']),
        ).count()
        expected = total_employees * _count_working_days(s_start, s_end, working_days)
        return round((present_ct / expected) * 100, 1) if expected else 0

    cur_start, cur_end = _session_bounds(date(year, month, 1))
    prev_start, prev_end = _session_bounds(date(cur_start.year - 1, 4, 1))
    session_comparison = [
        {'label': f'{prev_start.year}-{str(prev_end.year)[2:]}', 'value': _session_pct(prev_start, prev_end)},
        {'label': f'{cur_start.year}-{str(cur_end.year)[2:]}', 'value': _session_pct(cur_start, cur_end)},
    ]

    return {
        'kpis': kpis,
        'trend': {
            'attendance_percent': trend_attendance,
            'late_count': trend_late,
            'working_hours': trend_hours,
            'regularization_count': trend_reg,
        },
        'rankings': {'top_attendance': top_attendance, 'most_late': most_late, 'most_leave': most_leave},
        'heatmap': heatmap,
        'comparison': {'monthly': monthly_comparison, 'session': session_comparison},
    }


def _render_analytics_excel(payload):
    try:
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError('openpyxl not installed — run: pip install openpyxl')

    wb = Workbook()
    ws = wb.active
    ws.title = 'KPIs'
    ws.append(['Metric', 'Value'])
    for k, v in payload['kpis'].items():
        ws.append([k.replace('_', ' ').title(), v])

    ws2 = wb.create_sheet('Top Attendance')
    ws2.append(['Employee ID', 'Name', 'Role', 'Attendance %'])
    for r in payload['rankings']['top_attendance']:
        ws2.append([r['employee_id'], r['employee_name'], r['role'], r['attendance_percent']])

    ws3 = wb.create_sheet('Most Late')
    ws3.append(['Employee ID', 'Name', 'Role', 'Late Days'])
    for r in payload['rankings']['most_late']:
        ws3.append([r['employee_id'], r['employee_name'], r['role'], r['late_days']])

    ws4 = wb.create_sheet('Most Absent')
    ws4.append(['Employee ID', 'Name', 'Role', 'Absent Days'])
    for r in payload['rankings']['most_leave']:
        ws4.append([r['employee_id'], r['employee_name'], r['role'], r['leave_days']])

    ws5 = wb.create_sheet('Heatmap')
    ws5.append(['Date', 'Attendance %'])
    for d in payload['heatmap']:
        ws5.append([d['date'], d['attendance_percent']])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _render_analytics_pdf(payload, month, year):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except ImportError:
        raise RuntimeError('reportlab not installed — run: pip install reportlab')

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = height - 50

    c.setFont('Helvetica-Bold', 14)
    c.drawString(40, y, f'Attendance Analytics — {calendar.month_name[month]} {year}')
    y -= 30

    c.setFont('Helvetica-Bold', 11)
    c.drawString(40, y, 'KPIs')
    y -= 18
    c.setFont('Helvetica', 10)
    for k, v in payload['kpis'].items():
        c.drawString(50, y, f'{k.replace("_", " ").title()}: {v}')
        y -= 14
        if y < 60:
            c.showPage()
            y = height - 50

    y -= 10
    c.setFont('Helvetica-Bold', 11)
    c.drawString(40, y, 'Top Attendance')
    y -= 18
    c.setFont('Helvetica', 10)
    for r in payload['rankings']['top_attendance']:
        c.drawString(50, y, f"{r['employee_id']} - {r['employee_name']} - {r['attendance_percent']}%")
        y -= 14
        if y < 60:
            c.showPage()
            y = height - 50

    c.showPage()
    c.save()
    buf.seek(0)
    return buf

# ═══════════════════════════════════════════════════════════════════════
#  SETTINGS
# ═══════════════════════════════════════════════════════════════════════

@staff_attendance_bp.route('/settings', methods=['GET'])
@role_required('PRINCIPAL', 'HR')
def get_settings():
    user = _current_user()
    settings = StaffAttendanceSettings.get_or_create(user.school_id)
    return jsonify(settings.to_dict())


@staff_attendance_bp.route('/settings', methods=['PUT'])
@role_required('PRINCIPAL', 'HR')
def update_settings():
    user = _current_user()
    settings = StaffAttendanceSettings.get_or_create(user.school_id)
    data = request.get_json() or {}

    old_snapshot = settings.to_dict()

    fields = [
        'school_address', 'latitude', 'longitude', 'radius_meters',
        'school_start_time', 'school_end_time', 'grace_minutes',
        'half_day_after_minutes', 'late_after_minutes', 'overtime_after_minutes',
        'auto_checkout_time', 'approval_required', 'mock_location_detection',
        'device_restriction', 'payroll_sync_enabled', 'attendance_cutoff_day',
    ]
    for f in fields:
        if f in data:
            setattr(settings, f, data[f])

    if 'working_days' in data and isinstance(data['working_days'], list):
        settings.working_days = ','.join(data['working_days'])

    if 'attendance_lock_date' in data:
        settings.attendance_lock_date = _parse_date(data['attendance_lock_date'])

    settings.updated_by = user.id
    db.session.commit()

    svc.log_audit(user.school_id, 'SETTINGS_CHANGED', action_by=user.id,
                  old_value=old_snapshot, new_value=settings.to_dict())
    db.session.commit()

    return jsonify(settings.to_dict())


# ═══════════════════════════════════════════════════════════════════════
#  CHECK-IN / CHECK-OUT  (teacher/staff self-service)
# ═══════════════════════════════════════════════════════════════════════

@staff_attendance_bp.route('/check-in', methods=['POST'])
@jwt_required()
def do_check_in():
    user = _current_user()
    data = request.get_json() or {}
    try:
        record = svc.check_in(
            user,
            latitude=data.get('latitude'),
            longitude=data.get('longitude'),
            accuracy=data.get('accuracy'),
            is_mock=data.get('is_mock', False),
            device=data.get('device'),
        )
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    return jsonify(record.to_dict()), 201


@staff_attendance_bp.route('/check-out', methods=['POST'])
@jwt_required()
def do_check_out():
    user = _current_user()
    data = request.get_json() or {}
    try:
        record = svc.check_out(user, latitude=data.get('latitude'), longitude=data.get('longitude'))
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    return jsonify(record.to_dict())


@staff_attendance_bp.route('/my-status', methods=['GET'])
@jwt_required()
def my_today_status():
    user = _current_user()
    record = StaffAttendance.query.filter_by(user_id=user.id, attendance_date=date.today()).first()
    return jsonify(record.to_dict() if record else None)


# ═══════════════════════════════════════════════════════════════════════
#  DASHBOARD  (top cards)
# ═══════════════════════════════════════════════════════════════════════

@staff_attendance_bp.route('/dashboard', methods=['GET'])
@role_required('PRINCIPAL', 'HR')
def dashboard():
    user = _current_user()
    on_date = _parse_date(request.args.get('date'), date.today())

    total_employees = User.query.filter(
        User.school_id == user.school_id,
        User.role != UserRole.STUDENT, User.role != UserRole.PARENT,
        User.is_active == True,
    ).count()

    rows = StaffAttendance.query.filter_by(school_id=user.school_id, attendance_date=on_date).all()

    present = sum(1 for r in rows if r.status == 'PRESENT')
    late    = sum(1 for r in rows if r.status == 'LATE')
    half    = sum(1 for r in rows if r.status == 'HALF_DAY')
    absent_marked = sum(1 for r in rows if r.status == 'ABSENT')
    pending = sum(1 for r in rows if r.approval_status == 'PENDING')
    marked_users = {r.user_id for r in rows}
    not_marked = total_employees - len(marked_users)
    absent_total = absent_marked + max(0, not_marked)

    reg_pending = StaffAttendanceRegularization.query.filter_by(
        school_id=user.school_id, status='PENDING'
    ).count()

    checkins = [r.check_in_time for r in rows if r.check_in_time]
    avg_checkin = None
    if checkins:
        avg_seconds = sum(c.hour * 3600 + c.minute * 60 + c.second for c in checkins) / len(checkins)
        avg_checkin = f'{int(avg_seconds // 3600):02d}:{int((avg_seconds % 3600) // 60):02d}'

    total_minutes = sum(r.working_minutes or 0 for r in rows if r.check_out_time)
    checked_out_count = sum(1 for r in rows if r.check_out_time)
    avg_working_hours = round((total_minutes / checked_out_count) / 60, 2) if checked_out_count else 0

    attendance_pct = round(((present + late + half) / total_employees) * 100, 2) if total_employees else 0

    return jsonify({
        'date': on_date.isoformat(),
        'total_employees': total_employees,
        'present_today': present,
        'absent_today': absent_total,
        'late_today': late,
        'half_day_today': half,
        'on_leave_today': 0,   # wired once Leave module integration lands
        'pending_approval': pending,
        'regularization_requests': reg_pending,
        'average_check_in': avg_checkin,
        'average_working_hours': avg_working_hours,
        'attendance_percent': attendance_pct,
    })


# ═══════════════════════════════════════════════════════════════════════
#  TODAY ATTENDANCE LIST  (card click -> filtered list)
# ═══════════════════════════════════════════════════════════════════════

@staff_attendance_bp.route('/today', methods=['GET'])
@role_required('PRINCIPAL', 'HR')
def today_list():
    user = _current_user()
    on_date = _parse_date(request.args.get('date'), date.today())
    status_filter    = request.args.get('status')            # PRESENT/LATE/ABSENT/HALF_DAY/MISSING_CHECKOUT
    approval_filter  = request.args.get('approval_status')   # PENDING/APPROVED/REJECTED
    search           = request.args.get('search', '').strip()
    role_filter      = request.args.get('role')
    page     = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))

    q = StaffAttendance.query.filter_by(school_id=user.school_id, attendance_date=on_date)
    if status_filter:
        q = q.filter(StaffAttendance.status == status_filter)
    if approval_filter:
        q = q.filter(StaffAttendance.approval_status == approval_filter)

    q = q.join(User, StaffAttendance.user_id == User.id)
    if role_filter:
        q = q.filter(User.role == role_filter)
    if search:
        like = f'%{search}%'
        q = q.filter(db.or_(User.name.ilike(like), User.employee_id.ilike(like)))

    total = q.count()
    rows = q.order_by(StaffAttendance.check_in_time.desc()).offset((page - 1) * per_page).limit(per_page).all()

    return jsonify({
        'total': total, 'page': page, 'per_page': per_page,
        'items': [r.to_dict() for r in rows],
    })


# ═══════════════════════════════════════════════════════════════════════
#  APPROVAL
# ═══════════════════════════════════════════════════════════════════════

@staff_attendance_bp.route('/approve/<int:record_id>', methods=['POST'])
@role_required('PRINCIPAL', 'HR')
def approve_one(record_id):
    approver = _current_user()
    record = StaffAttendance.query.get_or_404(record_id)
    if record.school_id != approver.school_id:
        return jsonify({'error': 'Not found'}), 404
    reason = (request.get_json() or {}).get('reason')
    svc.approve_attendance(record, approver, reason=reason)
    return jsonify(record.to_dict())


@staff_attendance_bp.route('/reject/<int:record_id>', methods=['POST'])
@role_required('PRINCIPAL', 'HR')
def reject_one(record_id):
    approver = _current_user()
    record = StaffAttendance.query.get_or_404(record_id)
    if record.school_id != approver.school_id:
        return jsonify({'error': 'Not found'}), 404
    reason = (request.get_json() or {}).get('reason')
    svc.reject_attendance(record, approver, reason=reason)
    return jsonify(record.to_dict())


@staff_attendance_bp.route('/approve-bulk', methods=['POST'])
@role_required('PRINCIPAL', 'HR')
def approve_bulk():
    approver = _current_user()
    data = request.get_json() or {}
    ids = data.get('ids', [])
    if not ids:
        # "Approve All" — every PENDING row for today at this school
        on_date = _parse_date(data.get('date'), date.today())
        ids = [r.id for r in StaffAttendance.query.filter_by(
            school_id=approver.school_id, attendance_date=on_date, approval_status='PENDING'
        ).all()]
    records = svc.bulk_approve(ids, approver, reason=data.get('reason'))
    return jsonify({'updated': len(records)})


@staff_attendance_bp.route('/reject-bulk', methods=['POST'])
@role_required('PRINCIPAL', 'HR')
def reject_bulk():
    approver = _current_user()
    data = request.get_json() or {}
    ids = data.get('ids', [])
    if not ids:
        on_date = _parse_date(data.get('date'), date.today())
        ids = [r.id for r in StaffAttendance.query.filter_by(
            school_id=approver.school_id, attendance_date=on_date, approval_status='PENDING'
        ).all()]
    records = svc.bulk_reject(ids, approver, reason=data.get('reason'))
    return jsonify({'updated': len(records)})


# ═══════════════════════════════════════════════════════════════════════
#  REGULARIZATION
# ═══════════════════════════════════════════════════════════════════════

@staff_attendance_bp.route('/regularization', methods=['POST'])
@jwt_required()
def create_regularization():
    user = _current_user()
    data = request.get_json() or {}
    att_date = _parse_date(data.get('date'), date.today())

    def _parse_dt(v):
        return datetime.fromisoformat(v) if v else None

    reg = svc.submit_regularization(
        user, att_date,
        reason_type=data.get('reason_type'),
        reason_text=data.get('reason_text'),
        requested_check_in=_parse_dt(data.get('requested_check_in')),
        requested_check_out=_parse_dt(data.get('requested_check_out')),
    )
    return jsonify(reg.to_dict()), 201


@staff_attendance_bp.route('/regularization', methods=['GET'])
@role_required('PRINCIPAL', 'HR')
def list_regularization():
    user = _current_user()
    status_filter = request.args.get('status')
    q = StaffAttendanceRegularization.query.filter_by(school_id=user.school_id)
    if status_filter:
        q = q.filter_by(status=status_filter)
    rows = q.order_by(StaffAttendanceRegularization.created_at.desc()).all()
    return jsonify([r.to_dict() for r in rows])


@staff_attendance_bp.route('/regularization/<int:reg_id>/review', methods=['POST'])
@role_required('PRINCIPAL', 'HR')
def review_regularization(reg_id):
    approver = _current_user()
    reg = StaffAttendanceRegularization.query.get_or_404(reg_id)
    if reg.school_id != approver.school_id:
        return jsonify({'error': 'Not found'}), 404
    data = request.get_json() or {}
    approve = bool(data.get('approve'))
    updated = svc.review_regularization(reg, approver, approve, review_note=data.get('review_note'))
    return jsonify(updated.to_dict())


# ═══════════════════════════════════════════════════════════════════════
#  MONTHLY SUMMARY
# ═══════════════════════════════════════════════════════════════════════

@staff_attendance_bp.route('/monthly-summary', methods=['GET'])
@role_required('PRINCIPAL', 'HR')
def monthly_summary():
    user = _current_user()
    month = int(request.args.get('month', date.today().month))
    year  = int(request.args.get('year', date.today().year))

    summaries = svc.rebuild_monthly_summary_for_school(user.school_id, month, year)
    return jsonify([s.to_dict() for s in summaries])


@staff_attendance_bp.route('/monthly-summary/<int:target_user_id>', methods=['GET'])
@jwt_required()
def monthly_summary_one(target_user_id):
    requester = _current_user()
    target = User.query.get_or_404(target_user_id)
    if target.school_id != requester.school_id:
        return jsonify({'error': 'Not found'}), 404
    # Self, or Principal/HR viewing anyone
    if requester.id != target.id and requester.role.value not in ('PRINCIPAL', 'HR', 'DIRECTOR', 'VICE_PRINCIPAL'):
        return jsonify({'error': 'Access denied'}), 403

    month = int(request.args.get('month', date.today().month))
    year  = int(request.args.get('year', date.today().year))
    summary = svc.rebuild_monthly_summary(requester.school_id, target, month, year)
    return jsonify(summary.to_dict())


# ═══════════════════════════════════════════════════════════════════════
#  EMPLOYEE ATTENDANCE PROFILE
# ═══════════════════════════════════════════════════════════════════════

@staff_attendance_bp.route('/employee/<int:target_user_id>/history', methods=['GET'])
@role_required('PRINCIPAL', 'HR')
def employee_history(target_user_id):
    user = _current_user()
    target = User.query.get_or_404(target_user_id)
    if target.school_id != user.school_id:
        return jsonify({'error': 'Not found'}), 404

    from_date = _parse_date(request.args.get('from_date'))
    to_date   = _parse_date(request.args.get('to_date'), date.today())

    q = StaffAttendance.query.filter_by(user_id=target.id)
    if from_date:
        q = q.filter(StaffAttendance.attendance_date >= from_date)
    q = q.filter(StaffAttendance.attendance_date <= to_date)
    rows = q.order_by(StaffAttendance.attendance_date.desc()).all()

    reg_rows = StaffAttendanceRegularization.query.filter_by(user_id=target.id).order_by(
        StaffAttendanceRegularization.created_at.desc()
    ).limit(50).all()

    return jsonify({
        'employee': {
            'user_id': target.id,
            'employee_id': target.employee_id,
            'name': target.name,
            'role': target.role.value,
            'designation': target.designation,
            'photo_url': target.avatar_url,
        },
        'daily_history': [r.to_dict() for r in rows],
        'regularization_history': [r.to_dict() for r in reg_rows],
    })


# ═══════════════════════════════════════════════════════════════════════
#  EMPLOYEE LIST  (used by Add-Staff-style pages; ensures employee_id exists)
# ═══════════════════════════════════════════════════════════════════════

@staff_attendance_bp.route('/employees', methods=['GET'])
@role_required('PRINCIPAL', 'HR')
def list_employees():
    user = _current_user()
    employees = User.query.filter(
        User.school_id == user.school_id,
        User.role != UserRole.STUDENT, User.role != UserRole.PARENT,
    ).order_by(User.name).all()

    changed = False
    for e in employees:
        if not e.employee_id:
            e.employee_id = generate_employee_id(user.school_id)
            changed = True
    if changed:
        db.session.commit()

    return jsonify([{
        'user_id': e.id, 'employee_id': e.employee_id, 'name': e.name,
        'role': e.role.value, 'designation': e.designation, 'photo_url': e.avatar_url,
        'is_active': e.is_active,
    } for e in employees])


# ═══════════════════════════════════════════════════════════════════════
#  ANALYTICS
# ═══════════════════════════════════════════════════════════════════════

@staff_attendance_bp.route('/analytics', methods=['GET'])
@role_required('PRINCIPAL', 'HR')
def analytics():
    user = _current_user()
    range_key = request.args.get('range', 'monthly')
    month = int(request.args.get('month', date.today().month))
    year = int(request.args.get('year', date.today().year))
    role_filter = request.args.get('role', 'ALL')

    payload = _analytics_payload(user.school_id, range_key, month, year, role_filter)
    return jsonify(payload)


@staff_attendance_bp.route('/analytics/export', methods=['GET'])
@role_required('PRINCIPAL', 'HR')
def analytics_export():
    user = _current_user()
    range_key = request.args.get('range', 'monthly')
    month = int(request.args.get('month', date.today().month))
    year = int(request.args.get('year', date.today().year))
    role_filter = request.args.get('role', 'ALL')
    fmt = request.args.get('format', 'excel')

    payload = _analytics_payload(user.school_id, range_key, month, year, role_filter)

    try:
        if fmt == 'pdf':
            buf = _render_analytics_pdf(payload, month, year)
            return send_file(buf, mimetype='application/pdf', as_attachment=True,
                              download_name=f'attendance-analytics-{month}-{year}.pdf')
        buf = _render_analytics_excel(payload)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                          as_attachment=True, download_name=f'attendance-analytics-{month}-{year}.xlsx')
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500
